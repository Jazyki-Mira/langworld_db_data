import csv
from collections import Counter
from collections.abc import Generator
from contextlib import contextmanager
from copy import deepcopy
from pathlib import Path
from typing import Any, Iterable, Literal, NamedTuple, Optional, Union

import _csv  # for typing only

# for typing:
import openpyxl.cell
import openpyxl.worksheet.worksheet
from openpyxl import load_workbook

CSVDelimiter = Literal[",", ";"]


def append_empty_column_to_csv(
    path_to_file: Path,
    name_of_new_column: str,
    delimiter: CSVDelimiter = ",",
    custom_path_to_output_file: Optional[Path] = None,
) -> None:
    """Adds empty column (as last column) to CSV file. **Overwrites file**,
    but optional output path can be specified to create a new file.

    :raises ValueError if column name already exists in file.
    :raises FileExistsError if custom output file is specified and already exists.
    """
    if custom_path_to_output_file is not None and custom_path_to_output_file.exists():
        raise FileExistsError(
            f"You provided a custom path to output file {custom_path_to_output_file},"
            " but it already exists. To append column in place, do not indicate custom"
            " output path."
        )
    output_path = custom_path_to_output_file or path_to_file

    rows = read_plain_rows_from_csv(path_to_file=path_to_file, delimiter=delimiter)
    header_row, data_rows = rows[0], rows[1:]

    if name_of_new_column in header_row:
        raise ValueError(f"Column {name_of_new_column} already exists in {path_to_file.name}")

    new_header_row: list[str] = header_row + [name_of_new_column]
    new_data_rows = [row + [""] for row in data_rows]

    write_csv(
        [new_header_row] + new_data_rows,
        path_to_file=output_path,
        overwrite=True,
        delimiter=delimiter,
    )


def check_csv_for_malformed_rows(path_to_file: Path) -> None:
    """
    Checks whether all rows in CSV file have the same number of columns.
    Throws IndexError if they do not.
    """
    rows = read_plain_rows_from_csv(path_to_file)
    row_count_for_number_of_columns = Counter(len(row) for row in rows)

    if len(row_count_for_number_of_columns) == 1:
        return

    # See what count is least frequent.  It is most likely that the least frequent
    # number of columns indicates a mistake. Although it is theoretically possible that
    # there are so many wrong rows in a file that the wrong number becomes more
    # frequent, it is very unlikely.
    least_frequent_numbers_of_columns = [
        item
        for item in row_count_for_number_of_columns
        if row_count_for_number_of_columns[item]
        == sorted(row_count_for_number_of_columns.values())[0]
    ]
    # I made it a list because it is theoretically possible that one row has one wrong
    # number of columns and one more row has one more wrong number of columns
    # (also wrong, but different)

    indices_of_likely_invalid_rows = []
    for i, row in enumerate(rows, start=1):
        if len(row) in least_frequent_numbers_of_columns:
            indices_of_likely_invalid_rows.append(str(i))

    raise IndexError(
        f"File {path_to_file.name}: Following rows have abnormal number of columns: "
        f"{', '.join(indices_of_likely_invalid_rows)}"
    )


def check_csv_for_repetitions_in_column(path_to_file: Path, column_name: str) -> None:
    """Throws ValueError if there are repetitions in given column of given file."""
    rows = read_dicts_from_csv(path_to_file)

    if column_name not in rows[0]:
        raise KeyError(
            f"Cannot check uniqueness of value in column <{column_name}> because it"
            " does not exist"
        )

    values_in_column = [row[column_name] for row in rows]

    counter = Counter(values_in_column)

    non_unique_keys = [key for key in counter if counter[key] > 1]

    if non_unique_keys:
        raise ValueError(
            f"File {path_to_file} has repeating values in column <{column_name}>:"
            f" {', '.join(non_unique_keys)}"
        )


def convert_xls_to_csv(
    path_to_input_excel_file: Path,
    sheet_name: str,
    path_to_output_csv_file: Path,
    delimiter: CSVDelimiter = ",",
    overwrite: bool = True,
) -> None:
    if not overwrite and path_to_output_csv_file.exists():
        raise FileExistsError(f"File {path_to_output_csv_file} already exists")

    with _load_worksheet(path_to_workbook=path_to_input_excel_file, sheet_name=sheet_name) as ws:
        rows_to_write = [_get_cell_values(row) for row in ws.iter_rows()]

        write_csv(
            rows=rows_to_write,
            path_to_file=path_to_output_csv_file,
            overwrite=overwrite,
            delimiter=delimiter,
        )


def _get_cell_values(row: Iterable[openpyxl.cell.Cell]) -> list[str]:
    """Takes a row produced from `.iter_rows()` and
    returns list of values of cells.

    **Note**: `openpyxl` returns `None` as value of an empty cell.
    This function changes this behavior in that
    it returns **an empty string** as value of an empty cell.
    """
    cell_values = []

    for cell in row:
        if cell.value is None:
            cell_values.append("")
        else:
            cell_values.append(cell.value)

    return cell_values


@contextmanager
def _load_worksheet(
    path_to_workbook: Path, sheet_name: str
) -> Generator[openpyxl.worksheet.worksheet.Worksheet, None, None]:
    """Opens Excel workbook, returns worksheet,
    closes the workbook after operation is done.
    """
    wb = load_workbook(path_to_workbook, data_only=True)
    try:
        yield wb[sheet_name]
    finally:
        wb.close()


def read_column_from_csv(path_to_file: Path, column_name: str) -> list[str]:
    """Reads one column from CSV file. Column name is taken from the top row.
    :raises KeyError if no such column is find.
    """
    return [row[column_name] for row in read_dicts_from_csv(path_to_file)]


def read_dicts_from_csv(path_to_file: Path, delimiter: CSVDelimiter = ",") -> list[dict[str, str]]:
    """Opens CSV file and reads it as or list of dictionaries
    (top row is considered row with keys, each row is a dictionary
    with keys taken from top row).
    """
    with path_to_file.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh, delimiter=delimiter))


def read_dict_from_2_csv_columns(
    path_to_file: Path, key_col: str, val_col: str, delimiter: CSVDelimiter = ","
) -> dict[str, str]:
    """Reads CSV file, returns data of two columns: one as keys, the other as values."""
    if key_col == val_col:
        raise ValueError(f"You passed same name for both key and value columns ({key_col})")

    rows = read_plain_rows_from_csv(path_to_file, delimiter=delimiter)

    header_row, data_rows = rows[0], rows[1:]

    # for mypy typechecking only
    if not isinstance(header_row, list) or not isinstance(data_rows, list):
        raise TypeError

    if not (key_col in header_row and val_col in header_row):
        raise KeyError(f"Name of {key_col=} or {val_col=} not found in {header_row=}")

    counter = Counter(header_row)

    if counter[key_col] > 1 or counter[val_col] > 1:
        raise ValueError(
            f"One of columns ({key_col=} or {val_col=}) was encountered "
            f"more than once in {header_row=}"
        )

    key_index, val_index = header_row.index(key_col), header_row.index(val_col)

    key_column = [row[key_index] for row in data_rows]
    if len(set(key_column)) < len(key_column):
        raise ValueError(
            f"Values in column {key_col} are not unique. "
            "Using them as keys may lead to unpredictable behavior."
        )

    value_column = [row[val_index] for row in data_rows]

    return {k: v for k, v in zip(key_column, value_column)}


def read_dicts_from_xls(path_to_file: Path, sheet_name: str) -> list[dict[str, str]]:
    """Opens Excel file and reads it as or list of dictionaries
    (top row is considered row with keys, each row is a dictionary
    with keys taken from top row).

    Empty value cells will result in **empty strings** (not `None`)
    as dictionary values.
    """

    with _load_worksheet(path_to_workbook=path_to_file, sheet_name=sheet_name) as ws:
        rows = list(ws.iter_rows())

        keys = _get_cell_values(rows[0])
        data_rows = [_get_cell_values(row) for row in rows[1:]]

        dicts = [
            {keys[col]: cell_value for col, cell_value in enumerate(data_row)}
            for data_row in data_rows
        ]

        return dicts


def read_plain_rows_from_csv(
    path_to_file: Path, delimiter: CSVDelimiter = ",", remove_1st_row: bool = False
) -> list[list[str]]:
    """Opens CSV file and reads it as plain rows (list of lists)."""
    with path_to_file.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        rows = list(reader)

    if remove_1st_row:
        return rows[1:]
    return rows


def write_csv(
    # can't use Iterable because mypy says it's not indexable
    rows: Union[
        list[list[str]],
        list[tuple[str, str]],
        list[tuple[str, ...]],
        tuple[list[str], ...],
        tuple[tuple[str, ...], ...],
        tuple[tuple[str, str], ...],
        list[dict[str, str]],
        tuple[dict[str, str], ...],
        tuple[NamedTuple, ...],
        list[NamedTuple],
    ],
    path_to_file: Path,
    overwrite: bool,
    delimiter: CSVDelimiter,
) -> None:
    """Writes rows to CSV file. All rows (items of main list) must be
    of same type (all lists, all tuples, all dicts or all NamedTuples).

    If rows are lists or tuples, plain rows will be written.
    If rows are dicts or NamedTuples, header row will be automatically added.
    """

    if not overwrite and path_to_file.exists():
        raise FileExistsError(f"File {path_to_file} already exists")

    types_of_rows = set([type(row) for row in rows])
    if len(types_of_rows) > 1:
        # Strictly speaking, I should be able to write list of lists combined with
        # tuples or list of dicts combined with NamedTuples, but this is
        # overcomplicating and potentially unpredictable. I should not be doing these
        # things in calling code. So for sake of simplicity, this is justified.
        raise TypeError(
            f"Cannot write items of different types ({types_of_rows}) "
            "in the same set of rows. "
            "All items have to be either lists, dicts or NamedTuples"
        )

    # it is bad practice to change input data structure, so making a copy
    rows_to_write = rows[:]

    print(f"Writing CSV to file {path_to_file}")

    with path_to_file.open(mode="w+", encoding="utf-8", newline="") as fh:
        first_row = rows[0]
        # noinspection PyUnusedLocal, PyProtectedMember, PyUnresolvedReferences
        writer: Union[csv.DictWriter[Any], _csv._writer, None] = None
        if hasattr(first_row, "_asdict"):
            # NamedTuple cannot be used in `isinstance` statement, so I use `hasattr`.
            # I have to put this check first, because isinstance(item, tuple)
            # will evaluate to True for NamedTuple, which will lead to wrong behavior
            # (absence of header row)
            # noinspection PyProtectedMember
            writer = csv.DictWriter(
                fh,
                fieldnames=list(first_row._asdict().keys()),
                delimiter=delimiter,
            )
            # noinspection PyProtectedMember
            rows_to_write = [row._asdict() for row in rows]  # type: ignore
        elif isinstance(first_row, dict):
            writer = csv.DictWriter(fh, fieldnames=list(first_row.keys()), delimiter=delimiter)
        elif isinstance(first_row, (list, tuple)):
            writer = csv.writer(fh, delimiter=delimiter)
        else:
            raise TypeError(
                f"Each item of the list of rows is of type {type(first_row)}. "
                "Supported types are list, tuple, dict, NamedTuple."
            )

        if isinstance(writer, csv.DictWriter):
            print("Writing header")
            writer.writeheader()

        writer.writerows(rows_to_write)  # type: ignore
        print(f"Written {len(rows_to_write)} rows")


def remove_one_matching_row(
    rows: list[dict[str, str]],
    lookup_column: str,
    match_content: str,
) -> tuple[list[dict[str, str]], int]:
    """
    Remove exactly one row from a list of dictionaries where the specified column
    matches the given content.

    Return the modified list of dictionaries and the
    0-based index of the removed row.

    Important!
        Incoming list is deepcopied, which means that original
        dictionaries in the list are not changed, and new ones are returned.

    Args:
        rows: List of dictionaries representing rows
        lookup_column: Name of the column to search in
        match_content: Content to search for in the specified column

    Returns:
        tuple: (modified_rows, index_of_removed_row)

    Raises:
        TypeError: If match_content is not str or int
        KeyError:
            - If the specified lookup_column is not found in the rows
            - If no rows match the specified match_content in the lookup_column

    Note:
        - The function only removes the first matching row
        - For removing multiple rows, use remove_multiple_matching_rows
    """
    if lookup_column not in rows[0]:
        # TODO add test
        raise KeyError(f"{lookup_column=} not found. Cannot remove a row")

    if type(match_content) not in (int, str):
        raise TypeError(
            f"match_content must be of type <str> or <int>, <{type(match_content)}> was given."
        )

    index_of_row_to_remove: Union[int, None] = None

    copied_rows = deepcopy(rows)

    for i, row in enumerate(copied_rows):

        if row[lookup_column] == match_content:
            index_of_row_to_remove = i
            break

    if index_of_row_to_remove is None:
        raise KeyError(
            f"{match_content=} not found in column {lookup_column=}. Cannot remove a row"
        )

    return (
        copied_rows[:index_of_row_to_remove] + copied_rows[index_of_row_to_remove + 1 :],
        index_of_row_to_remove,
    )


def remove_multiple_matching_rows(
    rows: list[dict[str, str]],
    lookup_column: str,
    match_content: str,
) -> tuple[list[dict[str, str]], tuple[int]]:
    """
    Remove all rows from a list of dictionaries where the specified column matches
    the given content. Returns the modified list of dictionaries and a tuple containing
    the 0-based indices of the first and last removed rows.

    Important!
        Incoming list is deepcopied, which means that original
        dictionaries in the list are not changed, and new ones are returned.

    Args:
        rows: List of dictionaries representing rows
        lookup_column: Name of the column to search in
        match_content: Content to search for in the specified column

    Returns:
        tuple: (modified_rows, (first_removed_index, last_removed_index))

    Raises:
        TypeError: If match_content is not str or int
        KeyError:
            - If the specified lookup_column is not found in the rows
            - If no rows match the specified match_content in the lookup_column

    Note:
        - Indices are 0-based
        - For removing exactly one row, use remove_one_matching_row
    """

    if type(match_content) not in (int, str):
        raise TypeError(
            f"match_content must be of type <str> or <int>, <{type(match_content)}> was given."
        )

    if lookup_column not in rows[0]:
        # TODO add test
        raise KeyError(f"{lookup_column=} not found. Cannot remove rows")

    line_numbers_of_removed_rows = []

    copied_rows = deepcopy(rows)

    for i, row in enumerate(copied_rows):
        if row[lookup_column] == match_content:
            line_numbers_of_removed_rows.append(i)

    if len(line_numbers_of_removed_rows) == 0:
        raise KeyError(
            f"{match_content=} not found in column {lookup_column=}. Couldn't remove rows"
        )

    first_line_number = line_numbers_of_removed_rows[0]
    last_line_number = line_numbers_of_removed_rows[-1]

    return (
        copied_rows[:first_line_number] + copied_rows[last_line_number + 1 :],
        (first_line_number, last_line_number),
    )


def remove_matching_rows(
    rows: list[dict[str, str]],
    lookup_column: str,
    match_content: str,
) -> tuple[list[dict[str, str]], tuple[int]]:
    """
    Remove all rows from a list of dictionaries where the specified column
    matches the given content. There may be single relevant row or
    several rows which may form a sequence or be disctributed across
    the given list.

    Return the modified list of dictionaries and the
    0-based index of the removed row.

    Important!
        Incoming list is deepcopied, which means that original
        dictionaries in the list are not changed, and new ones are returned.

    Args:
        rows: List of dictionaries representing rows
        lookup_column: Name of the column to search in
        match_content: Content to search for in the specified column

    Returns:
        tuple: (modified_rows, line_numbers_of_removed_rows)

    Raises:
        TypeError: If match_content is not str or int
        KeyError:
            - If the specified lookup_column is not found in the rows
            - If no rows match the specified match_content in the lookup_column

    Note:
        - The function only removes the first matching row
        - For removing multiple rows, use remove_multiple_matching_rows
    """

    if type(match_content) not in (int, str):
        raise TypeError(
            f"match_content must be of type <str> or <int>, <{type(match_content)}> was given."
        )
    
    if lookup_column not in rows[0]:
        raise KeyError(f"Column <{lookup_column}> not found. Cannot remove a row")

    line_numbers_of_removed_rows = []

    copied_rows = deepcopy(rows)

    for i, row in enumerate(copied_rows):
        if row[lookup_column] == match_content:
            line_numbers_of_removed_rows.append(i)

    if len(line_numbers_of_removed_rows) == 0:
        raise KeyError(
            f"{match_content=} not found in column {lookup_column=}. Couldn't remove rows"
        )
    
    new_rows = []

    for i, row in enumerate(copied_rows):
        if i in line_numbers_of_removed_rows:
            continue

        new_rows.append(row)

    for row in new_rows:
        print(row)

    return (
        new_rows,
        tuple(line_numbers_of_removed_rows),
    )
