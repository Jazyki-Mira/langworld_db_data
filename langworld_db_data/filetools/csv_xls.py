from collections import Counter
import _csv  # for typing only
import csv
from pathlib import Path
from typing import Literal, Union

from openpyxl import load_workbook

CSVDelimiter = Literal[',', ';']


def check_csv_for_malformed_rows(path_to_file: Path) -> None:
    """
    Checks whether all rows in CSV file have the same number of columns.
    Throws IndexError if they do not.
    """
    rows = read_plain_rows_from_csv(path_to_file)
    row_count_for_number_of_columns = Counter(len(row) for row in rows)

    if len(row_count_for_number_of_columns) == 1:
        return

    # See what count is least frequent.  It is most likely that the least frequent number of columns
    # indicates a mistake. Although it is theoretically possible that there are so many wrong rows
    # in a file that the wrong number becomes more frequent, it is very unlikely.
    least_frequent_numbers_of_columns = [
        item for item in row_count_for_number_of_columns
        if row_count_for_number_of_columns[item] == sorted(row_count_for_number_of_columns.values())[0]
    ]
    # I made it a list because it is theoretically possible that one row has one wrong number of columns
    # and one more row has one more wrong number of columns (also wrong, but different)

    indices_of_likely_invalid_rows = []
    for i, row in enumerate(rows, start=1):
        if len(row) in least_frequent_numbers_of_columns:
            indices_of_likely_invalid_rows.append(str(i))

    raise IndexError(f'File {path_to_file.name}: Following rows have abnormal number of columns: '
                     f'{", ".join(indices_of_likely_invalid_rows)}')


def check_csv_for_repetitions_in_column(path_to_file: Path, column_name: str) -> None:
    """Throws ValueError if there are repetitions in given column of given file.
    """
    rows = read_dicts_from_csv(path_to_file)

    if column_name not in rows[0]:
        raise KeyError(f'Cannot check uniqueness of value in column <{column_name}> because it does not exist')

    values_in_column = [row[column_name] for row in rows]

    counter = Counter(values_in_column)

    non_unique_keys = [key for key in counter if counter[key] > 1]

    if non_unique_keys:
        raise ValueError(
            f'File {path_to_file} has repeating values in column <{column_name}>: {", ".join(non_unique_keys)}')


def convert_xls_to_csv(
    path_to_input_excel_file: Path,
    sheet_name: str,
    path_to_output_csv_file: Path,
    delimiter: CSVDelimiter = ',',
    overwrite: bool = True,
) -> None:
    if not overwrite and path_to_output_csv_file.exists():
        raise FileExistsError(f'File {path_to_output_csv_file} already exists')

    wb = load_workbook(path_to_input_excel_file, data_only=True)
    ws = wb[sheet_name]

    rows_to_write = []
    for row in ws.iter_rows():
        rows_to_write.append([cell.value for cell in row])

    write_csv(
        rows=rows_to_write,
        path_to_file=path_to_output_csv_file,
        overwrite=overwrite,
        delimiter=delimiter,
    )
    wb.close()


def read_column_from_csv(path_to_file: Path, column_name: str) -> list[str]:
    """Reads one column from CSV file. Column name is taken from the top row.
    :raises KeyError if no such column is find.
    """
    return [row[column_name] for row in read_dicts_from_csv(path_to_file)]


def read_dicts_from_csv(path_to_file: Path, delimiter: CSVDelimiter = ',') -> list[dict[str, str]]:
    """Opens CSV file and reads it as or list of dictionaries
    (top row is considered row with keys, each row is a dictionary
    with keys taken from top row).
    """
    with path_to_file.open(mode='r', encoding='utf-8-sig', newline='') as fh:
        return list(csv.DictReader(fh, delimiter=delimiter))


def read_dict_from_2_csv_columns(path_to_file: Path,
                                 key_col: str,
                                 val_col: str,
                                 delimiter: CSVDelimiter = ',') -> dict[str, str]:
    """Reads CSV file, returns data of two columns: one as keys, the other one as values.
    """
    if key_col == val_col:
        raise ValueError(f'You passed same name for both key and value columns ({key_col})')

    rows = read_plain_rows_from_csv(path_to_file, delimiter=delimiter)

    header_row, data_rows = rows[0], rows[1:]

    # for mypy typechecking only
    if not isinstance(header_row, list) or not isinstance(data_rows, list):
        raise TypeError

    if not (key_col in header_row and val_col in header_row):
        raise KeyError(f'Name of {key_col=} or {val_col=} not found in {header_row=}')

    counter = Counter(header_row)

    if counter[key_col] > 1 or counter[val_col] > 1:
        raise ValueError(f'One of columns ({key_col=} or {val_col=}) was encountered '
                         f'more than once in {header_row=}')

    key_index, val_index = header_row.index(key_col), header_row.index(val_col)

    key_column = [row[key_index] for row in data_rows]
    if len(set(key_column)) < len(key_column):
        raise ValueError(f'Values in column {key_col} are not unique. '
                         f'Using them as keys may lead to unpredictable behavior.')

    value_column = [row[val_index] for row in data_rows]

    return {k: v for k, v in zip(key_column, value_column)}


def read_plain_rows_from_csv(path_to_file: Path,
                             delimiter: CSVDelimiter = ',',
                             remove_1st_row: bool = False) -> list[list[str]]:
    """Opens CSV file and reads it as plain rows (list of lists)."""
    with path_to_file.open(mode='r', encoding='utf-8-sig', newline='') as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        rows = list(reader)

    if remove_1st_row:
        return rows[1:]
    return rows


def write_csv(rows: Union[list, tuple], path_to_file: Path, overwrite: bool, delimiter: CSVDelimiter) -> None:
    """Writes rows to CSV file. All rows (items of main list) must be
    of same type (all lists, all tuples, all dicts or all NamedTuples).

    If rows are lists or tuples, plain rows will be written.
    If rows are dicts or NamedTuples, header row will be automatically added.
    """

    if not overwrite and path_to_file.exists():
        raise FileExistsError(f'File {path_to_file} already exists')

    types_of_rows = set([type(row) for row in rows])
    if len(types_of_rows) > 1:
        # Strictly speaking, I should be able to write list of lists combined with tuples
        # or list of dicts combined with NamedTuples, but this is overcomplicating
        # and potentially unpredictable. I should not be doing these things in calling code.
        # So for sake of simplicity, this is justified.
        raise TypeError(f'Cannot write items of different types ({types_of_rows}) '
                        'in the same set of rows. '
                        'All items have to be either lists, dicts or NamedTuples')

    # it is bad practice to change input data structure, so making a copy
    rows_to_write = rows[:]

    print(f'Writing CSV to file {path_to_file}')

    with path_to_file.open(mode='w+', encoding='utf-8', newline='') as fh:
        # noinspection PyUnusedLocal, PyProtectedMember, PyUnresolvedReferences
        writer: Union[csv.DictWriter, _csv._writer, None] = None  # for mypy typechecking only
        if hasattr(rows[0], '_asdict'):
            # NamedTuple cannot be used in `isinstance` statement, so I use `hasattr`.
            # I have to put this check first, because isinstance(item, tuple)
            # will evaluate to True for NamedTuple, which will lead to wrong behavior
            # (absence of header row)
            # noinspection PyProtectedMember
            writer = csv.DictWriter(fh, fieldnames=list(rows[0]._asdict().keys()), delimiter=delimiter)
            # noinspection PyProtectedMember
            rows_to_write = [row._asdict() for row in rows]
        elif isinstance(rows[0], dict):
            writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()), delimiter=delimiter)
        elif isinstance(rows[0], (list, tuple)):
            writer = csv.writer(fh, delimiter=delimiter)
        else:
            raise TypeError(f'Each item of the list of rows is of type {type(rows[0])}. '
                            'Supported types are list, tuple, dict, NamedTuple.')

        if isinstance(writer, csv.DictWriter):
            print('Writing header')
            writer.writeheader()

        writer.writerows(rows_to_write)
        print(f'Written {len(rows_to_write)} rows')
